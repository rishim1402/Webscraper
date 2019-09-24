from pandas.compat import FileNotFoundError
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from bs4 import BeautifulSoup
import time
from tkinter import *
from tkinter import ttk
import tkinter
from tkinter import filedialog
import csv
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.chrome.options import Options
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import glob
import os
import math
import os
from pathlib import Path


def get_download_path():
    """Returns the default downloads path for linux or windows"""
    if os.name == 'nt':
        import winreg
        sub_key = r'SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders'
        downloads_guid = '{374DE290-123F-4565-9164-39C4925E467B}'
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, sub_key) as key:
            location = winreg.QueryValueEx(key, downloads_guid)[0]
        return location
    else:
        return os.path.join(os.path.expanduser('~'), 'downloads')


#d = get_download_path()
# print(str(d))

key = 0
count = 0
choices1 = {}
root = Tk()
completed = 0
file = ""

def selectdir():
    root.filename = filedialog.askdirectory()
    global file
    file = str(root.filename)
    print(root.filename)


def Main(key):
    global filedin
    filedin = root.filename + '/DIN_alldata_' + key1 + '.xlsx'
    print(filedin)
    loadDriver(key)
    loadInstitutes(driver, key)


def loadState(driver):
    # SuitFiledStateSummaryAction (Page 3)
    global count
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    k = []
    # Pull text from all instances of <a> tag within BodyText div
    global list_items
    list_items = soup.find_all('a')
    for i in list_items:
        k.append(i.contents[0])

    # print("check")
    for i in range(4, len(k) - 1):
        # updatelabel()
        print("    " + k[i])
        time.sleep(1)
        try:
            time.sleep(1)
            # FiledAccountSearchAction (Page 4)
            driver.find_element_by_partial_link_text(k[i]).click()
            time.sleep(1)
        except NoSuchElementException as exception:
            driver.refresh()
            i = i - 1
            continue
        driver.find_element_by_xpath('//*[@id="downloadReport"]/div').click()
        time.sleep(3)
        temp = []
        d = get_download_path()
        all_data = pd.DataFrame()
        new_df = pd.DataFrame()
        data_folder = Path(d)
        # for i in range(0,len(dk)):
        # ck = os.path.join(dk[i],'/')
        # print(ck)
        #print("1")
        #print(d)
        #print(type(d))
        ck = data_folder / 'suitFiledSearchReport_*.xls'
        #print("1",type(d), type(ck),ck)
        for f in glob.glob(str(ck)):
        #"C:\Users\nagot\Downloads\suitFiledSearchReport_*.xls"):
            df = pd.read_excel(f)
            all_data = all_data.append(df, ignore_index=True)
            for row in all_data.iterrows():
                index, data = row
                temp.append(data.tolist())
            temp1 = []
            temp2 = []
            final = []
            for i in range(1, len(temp)):
                dl = []
                temp1 = temp[i]
                r = isinstance(temp1[5], float)
                if not r:
                    d = temp1[5].split(',')
                    # print(d)
                    for x in d:
                        dl.append(x.split('--'))
                        # print(dl)
                    for t in dl:
                        if len(t) == 2:
                            for i in range(0, 5):
                                temp2.append(temp1[i])
                            temp2.append(t[0])
                            if t[1] == 'NA':
                                temp2.append(t[1])
                            # print(int(t[1]))
                            else:
                                try:
                                    temp2.append(int(t[1]))
                                except ValueError as Ve:
                                    temp2.append(t[1])
                            temp2.append(temp1[6])
                            final.append(temp2)
                            new_df = new_df.append(final, ignore_index=True)
                            temp2 = []
                            final = []
                else:
                    continue
            # print(final)
            # print("helloo")

            os.remove(f)

        # print(all_data)

        append_df_to_excel(filedin, new_df, startrow=count)
        count = count + len(new_df.index)
        all_data.iloc[0:0]
        new_df.iloc[0:0]
        # print("CHECK!!!")
        # Traverse back to the page of states/Union territory
        driver.execute_script("window.history.go(-1)")
    # Traverse back to the page of Institutes
    driver.execute_script("window.history.go(-1)")


def loadDriver(key):
    temp = int(key) - 1
    global driver
    options = Options()
    options.headless = True
    options.add_argument("window-size=1200,1100")
    driver = webdriver.Chrome('chromedriver.exe', options=options)
    #download_dir = TemporaryDirectory().name
    #os.mkdir(download_dir)

    # Send a command to tell chrome to download files in download_dir without
    # asking.
    driver.command_executor._commands["send_command"] = (
        "POST",
        '/session/$sessionId/chromium/send_command'
    )
    global d
    d = get_download_path()
    #print(d)
    #print("2")
    #print(type(d))
    params = {
        'cmd': 'Page.setDownloadBehavior',
        'params': {
            'behavior': 'allow',
            'downloadPath': d# r"C:\Users\nagot\PycharmProjects\tests"
            #"C:\Users\nagot\Downloads"
        }
    }
    driver.execute("send_command", params)

    driver.get('https://suit.cibil.com/')
    Select(driver.find_element_by_xpath('//*[@id="croreAccount"]')).select_by_value('2')
    Select(driver.find_element_by_xpath('//*[@id="quarterIdCrore"]')).select_by_index(str(temp))
    time.sleep(0.5)
    driver.find_element_by_xpath('/html/body/div/div[2]/div[1]/div[4]/form/div[1]/div[3]/div[4]/img').click()


def loadInstitutes(driver, key):
    html0 = driver.page_source
    soup0 = BeautifulSoup(html0, 'html.parser')
    k0 = []
    # Pull text from all instances of <a> tag within BodyText div
    list_items0 = soup0.find_all('a')
    for i0 in list_items0:
        k0.append(i0.contents[0])
    for i0 in range(4, len(k0) - 1):
        # progress.step(1)
        print(k0[i0])
        time.sleep(1)
        try:
            driver.find_element_by_partial_link_text(k0[i0]).click()
            time.sleep(1)
            driver.refresh()
            time.sleep(1)
        except NoSuchElementException as exception:
            i0 = i0 - 1
            driver.refresh()
            continue

        loadState(driver)
    print("Thank you!")
    driver.quit()


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, index=False, header=False,  **to_excel_kwargs)

    # save the workbook
    writer.save()


def get_dict():
    options = Options()
    options.headless = True
    driver1 = webdriver.Chrome('chromedriver.exe', options=options)
    driver1.get('https://suit.cibil.com/')
    try:
        a = Select(driver1.find_element_by_xpath('//*[@id="quarterIdCrore"]'))
    except NoSuchElementException as exception:
        get_dict()
    i = 0
    keys = []
    date = []
    for j in a.options:
        i = i + 1
        keys.append(i)
    keys = keys[1::]
    # print(keys)
    for i in keys:
        temp = '//*[@id="quarterIdCrore"]/option[' + str(i) + ']'
        try:
            date.append(driver1.find_element_by_xpath(temp).text)
            time.sleep(0.2)
        except NoSuchElementException as exception:
            continue
    # print(date)
    global choices1
    choices1 = dict(zip(date, keys))
    # print(choices1)


def get_key():
    global choices1
    get_dict()
    ch = list(choices1.keys())
    root.geometry("300x175")
    root.title("Web Scraping")
    L0 = Label(root, text="Director Identification Number System", fg='black')
    L0.pack()
    L7 = Label(root, text="Please make sure you have a stable connection.", fg='red')
    L7.pack()
    L12 = Label(root, text="Please select the folder: ", fg='black')
    L12.pack()
    B1 = Button(root, text="Browse...", command=selectdir)
    B1.pack()
    L1 = Label(root, text='Please select the required quadrant: ', fg='black')
    L1.pack()
    tkvar = StringVar(root)
    tkvar.set(ch[0])
    menu = OptionMenu(root, tkvar, *ch)
    menu.pack()

    def change_dropdown(*args):
        # print(tkvar.get())
        global key, key1
        key1 = tkvar.get()
        key = choices1.get(key1)

    # link function to change dropdown
    tkvar.trace('w', change_dropdown)

    def helloCallBack():
        # print(key)
        Main(key)

    # print("Hello Python")
    B = Button(root, text="Submit", command=helloCallBack, activebackground='blue')
    B.pack()
    root.mainloop()


get_key()
